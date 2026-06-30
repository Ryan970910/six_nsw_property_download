from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from .db import DEFAULT_PROPERTY_TABLE, PostgresConfig, upload_dat_rows_with_skip_report
from .transform import OUTPUT_COLUMNS
from .weekly_dat import (
    dedupe_rows,
    discover_weekly_zip_dates,
    download_weekly_zip,
    latest_weekly_zip_date,
    parse_weekly_zip,
    validate_output_schema,
)


logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class WeeklyRunResult:
    weeks: list[date]
    downloaded_files: int
    staged_rows: int
    inserted_rows: int
    skipped_rows: int
    skipped_report: Path | None = None
    week_results: list[dict[str, Any]] = field(default_factory=list)


def run_weekly_dat_upload(
    db_config: PostgresConfig,
    *,
    target_table: str = DEFAULT_PROPERTY_TABLE,
    weeks: list[date] | None = None,
    latest: bool = False,
    work_dir: Path,
    keep_zip: bool = True,
) -> WeeklyRunResult:
    if latest:
        weeks = [latest_weekly_zip_date()]
    elif not weeks:
        discovered = discover_weekly_zip_dates()
        if not discovered:
            raise RuntimeError("No weekly dates were found.")
        weeks = [discovered[-1]]

    work_dir.mkdir(parents=True, exist_ok=True)
    zip_dir = work_dir / "zip" if keep_zip else None
    report_dir = work_dir / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)

    logger.info("weekly_dat_pipeline_start weeks=%s target_table=%s work_dir=%s", [w.isoformat() for w in weeks], target_table, work_dir)

    total_files = 0
    total_staged = 0
    total_inserted = 0
    all_skipped: list[dict[str, Any]] = []
    week_results: list[dict[str, Any]] = []

    for week in weeks:
        logger.info("weekly_dat_week_start week=%s", week.isoformat())
        weekly_zip = download_weekly_zip(week, output_dir=zip_dir)
        parsed = parse_weekly_zip(weekly_zip)
        validate_output_schema(parsed.rows)
        unique_rows, duplicate_rows = dedupe_rows(parsed.rows)
        uploadable_rows, invalid_rows = split_uploadable_rows(unique_rows)

        upload_result = upload_dat_rows_with_skip_report(
            db_config,
            uploadable_rows,
            target_table=target_table,
        )
        skipped_rows = duplicate_rows + invalid_rows + upload_result.skipped_rows
        for row in skipped_rows:
            row.setdefault("week", week.isoformat())
        all_skipped.extend(skipped_rows)

        total_files += parsed.files
        total_staged += upload_result.copied_rows
        total_inserted += upload_result.inserted_rows
        week_info = {
            "week": week.isoformat(),
            "files": parsed.files,
            "parsed_rows": len(parsed.rows),
            "source_duplicate_rows": len(duplicate_rows),
            "invalid_rows": len(invalid_rows),
            "staged_rows": upload_result.copied_rows,
            "inserted_rows": upload_result.inserted_rows,
            "skipped_rows": len(skipped_rows),
        }
        week_results.append(week_info)
        logger.info("weekly_dat_week_complete %s", week_info)

    report_path = None
    if all_skipped:
        dates_part = "_".join(week.strftime("%Y%m%d") for week in weeks)
        report_path = report_dir / f"weekly_dat_skipped_{dates_part}.xlsx"
        write_skipped_rows_xlsx(all_skipped, report_path)
        logger.info("weekly_dat_skipped_report_written path=%s rows=%s", report_path, len(all_skipped))
    else:
        logger.info("weekly_dat_skipped_report_not_needed skipped_rows=0")

    logger.info(
        "weekly_dat_pipeline_complete weeks=%s files=%s staged_rows=%s inserted_rows=%s skipped_rows=%s",
        [w.isoformat() for w in weeks],
        total_files,
        total_staged,
        total_inserted,
        len(all_skipped),
    )
    return WeeklyRunResult(
        weeks=weeks,
        downloaded_files=total_files,
        staged_rows=total_staged,
        inserted_rows=total_inserted,
        skipped_rows=len(all_skipped),
        skipped_report=report_path,
        week_results=week_results,
    )


def write_skipped_rows_xlsx(rows: list[dict[str, Any]], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)
    headers = ["skip_reason", "week"] + [column for column in OUTPUT_COLUMNS if column != "id"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Skipped Rows"
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        sheet.append([serialize_excel_value(row.get(header)) for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in sheet[get_column_letter(index)]:
            if cell.value is not None:
                max_len = min(max(max_len, len(str(cell.value))), 60)
        sheet.column_dimensions[get_column_letter(index)].width = max_len + 2

    workbook.save(output_path)


def split_uploadable_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    uploadable: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    required_columns = ("url_property_id", "sale_date", "sale_price")
    for row in rows:
        missing = [column for column in required_columns if row.get(column) in (None, "")]
        if missing:
            skipped = dict(row)
            skipped["skip_reason"] = f"missing_required_column:{','.join(missing)}"
            invalid.append(skipped)
        else:
            uploadable.append(row)
    logger.info("weekly_dat_uploadable_split uploadable_rows=%s invalid_rows=%s", len(uploadable), len(invalid))
    return uploadable, invalid


def serialize_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value
